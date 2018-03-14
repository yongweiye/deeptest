#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 18-3-12 下午10:17
# @Author  : gonghuihui
# @File    : excel_study.py
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib

if __name__ == "__main__":
    #写
    wb = Workbook()

    ws = wb.active

    ws['A1'] = "开源优测"

    ws.append([1, 2, 3])

    import datetime
    ws['A2'] = datetime.datetime.now()

    wb.save("简单excle写示例.xlsx")

    # 读
    wr = load_workbook("简单excle写示例.xlsx")

    sheets = wr.get_sheet_names()
    print(sheets)

    for sh in sheets:
        print("读取工作薄名称： ", sh)

    ws = wr.get_sheet_by_name(sheets[0])

    A1 = ws['A1'].value
    print("A1单元格的值： ", A1)

    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)

    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    ebook_str = response.read().decode()

    ebook_dict = eval(ebook_str)

    wb1 = Workbook()
    ws1 = wb.active
    ws.append("书名", "作者", "描述", "出版社", "价格")

    # 写入书信息
    for book in ebook_dict:
        ws.append(book["title"],
                  ",".join(book["author"]),
                  book["summary"],
                  book["publiser"],
                  book["price"])

    wb1.save("ebook.xlsx")