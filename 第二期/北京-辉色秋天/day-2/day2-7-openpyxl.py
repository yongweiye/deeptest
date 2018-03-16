#-*-coding:utf-8-*-
import urllib.request
from openpyxl import Workbook
from openpyxl import load_workbook

def write_t():

    print("写excel示例")
    work = Workbook()
    work_state = work.active

    work_state['A1']="这是测试"
    work_state.append([1,2,3])

    import datetime

    work_state['A3']=datetime.datetime.now()

    work.save("简单excel示例.xlsx")

def read_t():
    print("读取示例")
    wb = load_workbook("简单excel示例.xlsx")
    sheets = wb.get_sheet_names()

    print("type(sheets) is ",type(sheets))

    for sheet in sheets:
        print("sheet name is ",sheet)

    ws = wb.get_sheet_by_name(sheets[0])

    A1 = ws['A1'].value
    print("A1 is ",A1)

    for index in ('A2', 'A3', 'B2',"C2"):
        print(index, "单元格的值：", ws[index].value)

def save_book():
    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    ebook_str = response.read().decode()
    ebook_dict = eval(ebook_str)

    wb = Workbook()

    ws = wb.active
    ws.append(["书名", "作者", "描述", "出版社", "价格"])

    for book in ebook_dict['books']:
        ws.append([book['title'],','.join(book['author']), book['summary'],book['publisher'], book['price']])

    wb.save("book.xlsx")


if __name__ == "__main__":
    #write_t()
    #read_t()
    save_book()
